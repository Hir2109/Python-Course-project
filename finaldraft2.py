import openpyxl
import datetime, subprocess
import os

# Function to validate date
def validateDate(date):
    parts =  date.split('/')
    if len(parts) != 3:  #To check if format contains 3 parts (day, month, year)
        return False
    day = int(parts[0])  #I'm converting thisday part to integer
    month = int(parts[1]) #I'm converting this month part to integer
    year = int(parts[2])  #I'm converting thisyear part to integer
    if day < 1 or day > 31 or month < 1 or month > 12 or year < 0:
        return False #to check if day, month, and year are within valid range or not
    return True

#To create a workbook
workbook = openpyxl.load_workbook("D:\expenseTracker.xlsx")
sheet = workbook.active
print("Existing workbook is successfully loaded.")


lastRow = sheet.max_row #If sheet is not empty,this will get your last row with data

#Now, if the sheet is empty,it'll add headers
if lastRow == 0:
    headers = ['Date', 'Category', 'Description', 'Amount', 'Total Expense']
    sheet.append(headers)  #This appends the headers to the sheet
    lastRow += 1 #to update lastRow counter
    
#paceholder for expenses list
expenses = []

#For user input number of expenses
numExpenses = None
while numExpenses is None:
    try:
        numExpenses = int(input("Please enter number of expenses: ")) # Convert input to integer
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        continue
    if numExpenses <= 0:
        print("Invalid input. Number of expenses should be greater than 0.")
        numExpenses = None

#For user input 
for i in range(numExpenses):
    date = input("Plese enter date (DD/MM/YYYY): ") #to get user input for date
    if validateDate(date): #to validate the date format
        print("Thanks!")
    else:
        print("Invalid date format.")
        date = input("Please enter date (DD/MM/YYYY): ") #asking user for input until they give correct format
        print("Thanks!")
        
    category = input("Please enter your category: ") #user input for category
    description = input("Please enter your description: ") #user input to get description
    amount = float(input("Please enter your amount: ")) #user input in float to get amount spent by them

    #To append expense data to list
    expenses.append({'date': date, 'category': category, 'description': description, 'amount': amount})

    #To write expense to spreadsheet
    sheet.cell(row=lastRow+i, column=1).value = date #to write date to cell in the first column
    sheet.cell(row=lastRow+i, column=2).value = category  #to write category to cell in the second column
    sheet.cell(row=lastRow+i, column=3).value = description #to write description to cell in the third column
    sheet.cell(row=lastRow+i, column=4).value = amount #to write amount to cell in the fourth column

#To calculate total expense of the day
totalExpense = {}
for expense in expenses: #This is a loop that iterates over each expense in the expenses list.
    date = expense['date']
    amount = expense['amount']
    if date in totalExpense:  #if the date already exists in totalExpense
        totalExpense[date] += amount #the amount is added to the existing value in the totalExpense dictionary
    else:
        totalExpense[date] = amount #adds a new key-value pair to the totalExpense

#To write total expense of the day to the sheet
for row in sheet.iter_rows(min_row=lastRow, min_col=1, max_row=lastRow+len(expenses)-1, max_col=5):
    date = row[0].value #to get value in first column
    if date in totalExpense:
        row[4].value = totalExpense[date] #totalexpense is written in fifth column (column index 4)

# Save the workbook
workbook.save("D:\expenseTracker.xlsx") #to save updated workbook
print("Expenses saved to expenseTracker.xlsx") #to print the message that data has been saved

for date, expense in totalExpense.items():
    if expense > 50: #to set the budget of $50
        notepadFile = open('notepad_message.txt', 'w') #opens notepad application to write data
        notepadFile.write(f"Total expenses for {date} was: ${expense:.2f}\n")#this message write total amount spent by user
        notepadFile.write(f"You have exceeded your budget by ${expense - 50.00:.2f}.")#this message gives extra amount spent over $50
        notepadFile.close() #to close notepad application
        os.system('start notepad notepad_message.txt')
       
# Function to validate week format
def validateWeek(week):
    try:
        startDate, endDate = week.split('-')
        startDate = startDate.strip()
        endDate = endDate.strip()
        datetime.datetime.strptime(startDate, '%d/%m/%Y') 
        datetime.datetime.strptime(endDate, '%d/%m/%Y')
        return True
    except ValueError:
        return False
    
#input for week   
calculateTotal = input("Do you want to calculate the total expenses for the week? (Y/N): ") 
if calculateTotal.lower() == 'y': #when users says yes to calculating total of the week
    week = input("Enter week (DD/MM/YYYY - DD/MM/YYYY): ")
    while not validateWeek(week):
        print("Invalid week format.")
        week = input("Enter week (DD/MM/YYYY - DD/MM/YYYY): ")

    startDate, endDate = week.split('-')
    startDate = startDate.strip()
    endDate = endDate.strip()

    totalExpenseWeek = 0
    for expense in expenses:
        date = expense['date']
        amount = expense['amount']
        if datetime.datetime.strptime(startDate, '%d/%m/%Y') <= datetime.datetime.strptime(date, '%d/%m/%Y') <= datetime.datetime.strptime(endDate, '%d/%m/%Y'):
            totalExpenseWeek += amount

    # Write total expense of the week to the notepad
     # Write total expense of the week to the notepad
    notepadFile = open('notepad_message.txt', 'a')
    notepadFile.write(f"\nTotal expenses for the week {startDate} - {endDate} was: ${totalExpenseWeek:.2f}\n")
    subprocess.Popen(['notepad', 'notepad_message.txt'])  #to open notepad with message using subprocess
    notepadFile.close() #to close notepad application
else:
    print("Thanks!") #if user said no it will just print thanks!"""

    
